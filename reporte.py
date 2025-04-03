import json
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

def generar_reporte_excel(json_file, excel_file):
    # Cargar datos desde el archivo JSON
    with open(json_file, 'r', encoding='utf-8') as file:
        datos = json.load(file)

    # Crear un libro de Excel y una hoja
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Registro"

    # Agregar título al reporte
    ws.merge_cells("A1:Y1")
    titulo = ws.cell(row=1, column=1, value="Reportes")  # Título del reporte
    titulo.font = Font(bold=True, size=14)
    titulo.alignment = Alignment(horizontal='center')  # Centrar el título

    # Definir bordes
    borde = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )

    # Definir encabezados con columna de correlativo
    encabezados = [
        "N°", "Código", "Apellidos y Nombres", "1", "2", "3", "4", "5", "6", "7", "8", "9", "10", "11", "12", "13", "14", "15", "16", "17", "18", "19", "20",
        "Prom E.P.", "E.F.", "Prom T.A.", "E.P. (30%)", "E.F. (50%)", "Prom Final"
    ]
    ws.append(encabezados)

    # Aplicar estilos a los encabezados
    for col_num, col_name in enumerate(encabezados, start=1):
        cell = ws.cell(row=2, column=col_num, value=col_name)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.border = borde  # Aplicar bordes

    # Llenar los datos en la hoja con el correlativo
    for index, persona in enumerate(datos, start=1):
        # Obtener los 20 días (si están presentes en el JSON)
        dias = persona.get("dias", ["F"] * 20)  # Rellenar con 'F' si no hay datos

        # Calcular los valores de las tareas académicas
        prom_ep = persona.get("prom_ep", "")
        ef = persona.get("ef", "")
        prom_ta = persona.get("prom_ta", "")
        ep_30 = prom_ep * 0.3 if prom_ep else ""
        ef_50 = ef * 0.5 if ef else ""
        prom_final = (ep_30 + ef_50) if ep_30 and ef_50 else ""

        # Agregar los datos de cada persona en su fila
        row = [
            index,  # Correlativo
            persona.get("codigo", ""),  # Código
            persona.get("nombre", ""),  # Apellidos y nombres
            *dias,  # Los 20 días (rellenar con valores de JSON o 'F')
            prom_ep,  # Prom E.P.
            ef,  # E.F.
            prom_ta,  # Prom T.A.
            ep_30,  # E.P. (30%)
            ef_50,  # E.F. (50%)
            prom_final  # Prom Final
        ]
        ws.append(row)

        # Aplicar bordes a cada celda
        for col_num in range(1, len(row) + 1):
            cell = ws.cell(row=index + 2, column=col_num)
            cell.border = borde  # Bordes en todas las celdas

        # Hacer que los números de la columna "N°" sean negros suaves
        ws.cell(row=index + 2, column=1).font = Font(size=10, bold=True, color="202020")  # Negro suave
        ws.cell(row=index + 2, column=1).alignment = Alignment(horizontal='center')  # Centrar correlativo

    # Ajustar el ancho de columnas automáticamente
    for col in ws.iter_cols(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        max_length = 0
        col_letter = col[0].column_letter  # Obtener la letra de la columna
        for cell in col:
            if cell.value and isinstance(cell, openpyxl.cell.cell.Cell):  # Evitar MergedCell
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    # Guardar el archivo Excel
    wb.save(excel_file)
    print(f"Reporte generado con éxito: {excel_file}")

# Ejemplo de uso
generar_reporte_excel("registro.json", "reporte_registro.xlsx")

#python reporte.py